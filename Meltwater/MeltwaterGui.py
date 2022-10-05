# <<<<<<< HEAD
import tkinter as tk
from tkinter import ttk
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import load_workbook


# Function that counts duplicate documents in Meltwater and converts them to integers
def DupCounter(d):
    if "Duplicate" in d:
        d = d.strip(" Duplicate")

    if "Duplicates" in d:
        d = d.strip(" Duplicates")

    return int(d) + 1


# Creating the tkinter window
root2 = tk.Tk()
root2.title("Login")
root2.geometry("300x100")
usernameLabel = ttk.Label(root2, text="User Name").grid(row=0, column=0)
un = tk.StringVar()
usernameEntry = ttk.Entry(root2, textvariable=un).grid(row=0, column=1)

# password label and password entry box
passwordLabel = ttk.Label(root2, text="Password").grid(row=1, column=0)
pw = tk.StringVar()
passwordEntry = ttk.Entry(root2, textvariable=pw, show='*').grid(row=1, column=1)


# Function to close the login window
def Close():
    root2.destroy()


# Button for closing
exit_button = ttk.Button(root2, text="Login", command=Close)
exit_button.grid(column=1, row=3, sticky=tk.E, padx=5, pady=5)

root2.mainloop()

# This allows for the program to interact with the webpage (The chromedriver.exe must be updated when a Google Chrome
#  browser update is released)
PATH = "C:\Python\ArticleCounter\chromedriver.exe"
driver = webdriver.Chrome(PATH)
# Logs into Meltwater Account
try:
    driver.get("https://app.meltwater.com/search/edit/8783647")
    time.sleep(4)
    user = driver.find_element_by_id('input_0')
    # "AFGlobalStrike@gmail.com"
    user.send_keys(tk.StringVar.get(un))
    time.sleep(4)
    user.send_keys(Keys.RETURN)
    time.sleep(4)
    password = driver.find_element_by_name('password')
    # SACback2020!
    password.send_keys(tk.StringVar.get(pw))
    password.send_keys(Keys.RETURN)
except selenium.common.exceptions.NoSuchElementException:
    time.sleep(5)
    driver.refresh()
    driver.get("https://app.meltwater.com/search/edit/8783647")
    time.sleep(4)
    user = driver.find_element_by_id('input_0')
    user.send_keys("AFGlobalStrike@gmail.com")
    time.sleep(4)
    user.send_keys(Keys.RETURN)
    time.sleep(7)
    password = driver.find_element_by_name('password')
    password.send_keys("SACback2020!")
    password.send_keys(Keys.RETURN)
# Waits for Meltwater to load webpage
print("Wait time")
wait = WebDriverWait(driver, 600)
wait.until(EC.visibility_of_element_located((By.XPATH,
                                             '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[25]/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger[1]/span')))

# Window for excel and document count
root = tk.Tk()
root.geometry("300x100")
root.title('Article Counter')
root.resizable(0, 0)

# configure the grid
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=3)

# document
document_label = ttk.Label(root, text="Enter number of documents:")
document_label.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)

document = tk.Entry(root)
document.grid(column=1, row=0, sticky=tk.E, padx=5, pady=5)

excel_label = ttk.Label(root, text="Excel file name:")
excel_label.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)

excel_entry = ttk.Entry(root)
excel_entry.grid(column=1, row=1, sticky=tk.E, padx=5, pady=5)


# This function gets the article count and sorts the article into the Excel sheet by the search parameters of the
# article
def getinfo():
    articleCount = 0
    docCounter = int(document.get()) + 1
    Fname = excel_entry.get() + ".xlsx"
    # Will need the document count so the program will not look for extra documents that do not exist
    # Input the Excel file name you want the information to be exported to
    wb = load_workbook(Fname)
    wsheet = wb.active
    names = wsheet['A']
    # Counts the articles and exports that number to the Excel file
    try:
        for i in range(1, int(docCounter)):
            print(i)
            texts = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    i) + ']/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger/span').text
            column = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    i) + ']/div/div/div[1]/div[2]/div/div/ng-transclude[6]/mw-content-document-keywords/div').text
            if "Reach" or "Social Echo" in texts:
                articleCount += 1
            else:
                Dup = DupCounter(texts)
                articleCount += Dup
            for x in range(0, len(names)):
                if column is None:
                    print("None Search")
                if column in names[x].value:
                    cell = 'B' + str(x + 1)
                    wsheet[cell].value += articleCount
                    wb.save(Fname)
                    articleCount = 0


    # If the counter tries to count a document that doesn't exist then it will output the number of articles counted
    # so far so the program will not crash, so you can run the program multiple times
    except selenium.common.exceptions.NoSuchElementException:
        articleCount += 1
        print("Article count was not completed but stopped at document number %s" % i)
        for f in range(i + 3, docCounter):
            print(f)
            texts = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    f) + ']/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger/span').text
            column = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    f) + ']/div/div/div[1]/div[2]/div/div/ng-transclude[6]/mw-content-document-keywords/div').text
            if "Reach" or "Social Echo" in texts:
                articleCount += 1
            else:
                Dup = DupCounter(texts)
                articleCount += Dup
            for x in range(0, len(names)):
                if column is None:
                    print("None Search")
                if column in names[x].value:
                    cell = 'B' + str(x + 1)
                    wsheet[cell].value += articleCount
                    wb.save(Fname)
                    articleCount = 0


    finally:
        print("Article Counter is done")


submit = tk.Button(root, text="Submit", command=getinfo)
submit.grid(column=1, row=3, sticky=tk.E, padx=5, pady=5)

root.mainloop()
# =======
import tkinter as tk
from tkinter import ttk
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import load_workbook


# Function that counts duplicate documents in Meltwater and converts them to integers
def DupCounter(d):
    if "Duplicate" in d:
        d = d.strip(" Duplicate")

    if "Duplicates" in d:
        d = d.strip(" Duplicates")

    return int(d) + 1


# Creating the tkinter window
root2 = tk.Tk()
root2.title("Login")
root2.geometry("300x100")
usernameLabel = ttk.Label(root2, text="User Name").grid(row=0, column=0)
un = tk.StringVar()
usernameEntry = ttk.Entry(root2, textvariable=un).grid(row=0, column=1)

# password label and password entry box
passwordLabel = ttk.Label(root2, text="Password").grid(row=1, column=0)
pw = tk.StringVar()
passwordEntry = ttk.Entry(root2, textvariable=pw, show='*').grid(row=1, column=1)


# Function to close the login window
def Close():
    root2.destroy()


# Button for closing
exit_button = ttk.Button(root2, text="Login", command=Close)
exit_button.grid(column=1, row=3, sticky=tk.E, padx=5, pady=5)

root2.mainloop()

# This allows for the program to interact with the webpage (The chromedriver.exe must be updated when a Google Chrome
#  browser update is released)
PATH = "C:\Python\ArticleCounter\chromedriver.exe"
driver = webdriver.Chrome(PATH)
# Logs into Meltwater Account
try:
    driver.get("https://app.meltwater.com/search/edit/8783647")
    time.sleep(4)
    user = driver.find_element_by_id('input_0')
    # "AFGlobalStrike@gmail.com"
    user.send_keys(tk.StringVar.get(un))
    time.sleep(4)
    user.send_keys(Keys.RETURN)
    time.sleep(4)
    password = driver.find_element_by_name('password')
    # SACback2020!
    password.send_keys(tk.StringVar.get(pw))
    password.send_keys(Keys.RETURN)
except selenium.common.exceptions.NoSuchElementException:
    time.sleep(5)
    driver.refresh()
    driver.get("https://app.meltwater.com/search/edit/8783647")
    time.sleep(4)
    user = driver.find_element_by_id('input_0')
    user.send_keys("AFGlobalStrike@gmail.com")
    time.sleep(4)
    user.send_keys(Keys.RETURN)
    time.sleep(7)
    password = driver.find_element_by_name('password')
    password.send_keys("SACback2020!")
    password.send_keys(Keys.RETURN)
# Waits for Meltwater to load webpage
print("Wait time")
wait = WebDriverWait(driver, 600)
wait.until(EC.visibility_of_element_located((By.XPATH,
                                             '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[25]/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger[1]/span')))

# Window for excel and document count
root = tk.Tk()
root.geometry("300x100")
root.title('Article Counter')
root.resizable(0, 0)

# configure the grid
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=3)

# document
document_label = ttk.Label(root, text="Enter number of documents:")
document_label.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)

document = tk.Entry(root)
document.grid(column=1, row=0, sticky=tk.E, padx=5, pady=5)

excel_label = ttk.Label(root, text="Excel file name:")
excel_label.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)

excel_entry = ttk.Entry(root)
excel_entry.grid(column=1, row=1, sticky=tk.E, padx=5, pady=5)


# This function gets the article count and sorts the article into the Excel sheet by the search parameters of the
# article
def getinfo():
    articleCount = 0
    docCounter = int(document.get()) + 1
    Fname = excel_entry.get() + ".xlsx"
    # Will need the document count so the program will not look for extra documents that do not exist
    # Input the Excel file name you want the information to be exported to
    wb = load_workbook(Fname)
    wsheet = wb.active
    names = wsheet['A']
    # Counts the articles and exports that number to the Excel file
    try:
        for i in range(1, int(docCounter)):
            print(i)
            texts = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    i) + ']/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger/span').text
            column = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    i) + ']/div/div/div[1]/div[2]/div/div/ng-transclude[6]/mw-content-document-keywords/div').text
            if "Reach" or "Social Echo" in texts:
                articleCount += 1
            else:
                Dup = DupCounter(texts)
                articleCount += Dup
            for x in range(0, len(names)):
                if column is None:
                    print("None Search")
                if column in names[x].value:
                    cell = 'B' + str(x + 1)
                    wsheet[cell].value += articleCount
                    wb.save(Fname)
                    articleCount = 0


    # If the counter tries to count a document that doesn't exist then it will output the number of articles counted
    # so far so the program will not crash, so you can run the program multiple times
    except selenium.common.exceptions.NoSuchElementException:
        articleCount += 1
        print("Article count was not completed but stopped at document number %s" % i)
        for f in range(i + 3, docCounter):
            print(f)
            texts = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    f) + ']/div/div/div[1]/div[2]/ng-transclude[2]/mw-content-document-footer/div/ng-transclude[1]/mw-content-document-action-trigger/span').text
            column = driver.find_element_by_xpath(
                '//*[@id="infiniteScrollParent"]/content-stream-container/section/section/md-card/mw-content-stream/div/ng-transclude[3]/mw-content-stream-body/div[1]/mw-media-document[' + str(
                    f) + ']/div/div/div[1]/div[2]/div/div/ng-transclude[6]/mw-content-document-keywords/div').text
            if "Reach" or "Social Echo" in texts:
                articleCount += 1
            else:
                Dup = DupCounter(texts)
                articleCount += Dup
            for x in range(0, len(names)):
                if column is None:
                    print("None Search")
                if column in names[x].value:
                    cell = 'B' + str(x + 1)
                    wsheet[cell].value += articleCount
                    wb.save(Fname)
                    articleCount = 0


    finally:
        print("Article Counter is done")


submit = tk.Button(root, text="Submit", command=getinfo)
submit.grid(column=1, row=3, sticky=tk.E, padx=5, pady=5)

root.mainloop()
# >>>>>>> afab0d05ffe23591d2d09ccb1328087ac01fa38c
